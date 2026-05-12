import csv
import datetime
from sqlite3 import IntegrityError
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from .models import Barrio, DetalleRuta, RutaBarrio, Tiempo, TipoUsuario, Usuario,Ruta, Empresa, Horario, Ruta
from .forms import ActualizarCuentaForm, BarrioForm, BusquedaBarriosForm, DetalleRutaForm, EmpresaForm, HorarioForm, LoginForm, RegistroUsuarioForm, RutaForm, TiempoForm
from django.contrib import messages
from functools import wraps
from django.db.models import Sum, Avg
from django.http import HttpResponse
from datetime import datetime
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Sum, Avg, Count
from io import BytesIO
import json
from django.db.models.functions import TruncDate
import openpyxl

def login_view(request):
    error = None
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            contrasena = form.cleaned_data['contrasena']
            try:
                usuario = Usuario.objects.get(email=email, contrasena=contrasena)
                request.session['usuario_id'] = usuario.idUsuario
                request.session['tipo'] = 'Administrador' if usuario.idTipoUsuario_id == 1 else 'Usuario general'

                if usuario.idTipoUsuario_id == 1:
                    return redirect('admin_dashboard')
                else:
                    return redirect('usuario_dashboard')
            except Usuario.DoesNotExist:
                error = "Credenciales inválidas"
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form, 'error': error})

def usuario_autenticado(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.session.get('usuario_id'):
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper

def logout_view(request):
    request.session.flush()
    return redirect('login')

def admin_dashboard(request):
    if not request.session.get('usuario_id'):
        return redirect('login')
    return render(request, 'admin_dashboard.html')

def usuario_dashboard(request):
    empresas = Empresa.objects.all()
    rutas = Ruta.objects.all()
    horarios = Horario.objects.all()
    barrios = Barrio.objects.all()

    # Leer parámetros GET para filtro barrios
    barrios_seleccionados_ids = request.GET.getlist('barrios')
    barrio_query = bool(barrios_seleccionados_ids)

    if barrio_query:
        barrios_filtrados = barrios.filter(pk__in=barrios_seleccionados_ids)
        rutas = rutas.filter(rutabarrio__idBarrio__pk__in=barrios_seleccionados_ids).distinct()

    else:
        barrios_filtrados = barrios

    # Crea el formulario con los datos GET actuales
    form = BusquedaBarriosForm(request.GET or None)
    form.fields['barrios'].choices = [(b.idBarrio, b.nombreBarrio) for b in barrios]


    context = {
        'empresas': empresas,
        'rutas': rutas,
        'horarios': horarios,
        'barrios': barrios_filtrados,
        'barrio_query': barrio_query,
        'barrios_seleccionados': barrios_seleccionados_ids,
        'form': form,  # ¡Aquí está el formulario!
    }

    return render(request, 'usuario_dashboard.html', context)

def filtrar_rutas_por_barrios(request):
    ids = request.GET.getlist('barrios[]')
    rutas_filtradas = Ruta.objects.filter(barrios__id__in=ids).distinct()
    return render(request, 'partials/rutas_filtradas.html', {'rutas': rutas_filtradas})

def lista_empresas(request):
    empresas = Empresa.objects.all()
    return render(request, 'empresas/lista.html', {'empresas': empresas})

def crear_empresa(request):
    if request.method == 'POST':
        form = EmpresaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_empresas')
    else:
        form = EmpresaForm()
    return render(request, 'empresas/formulario.html', {'form': form})

def editar_empresa(request, id):
    empresa = get_object_or_404(Empresa, idEmpresa=id)
    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            return redirect('lista_empresas')
    else:
        form = EmpresaForm(instance=empresa)
    return render(request, 'empresas/formulario.html', {'form': form, 'accion': 'Editar'})

def eliminar_empresa(request, id):
    empresa = get_object_or_404(Empresa, idEmpresa=id)
    if request.method == 'POST':
        empresa.delete()
        return redirect('lista_empresas')
    return render(request, 'empresas/confirmar_eliminar.html', {'empresa': empresa})

def lista_usuarios(request):
    if not request.session.get('usuario_id'):
        return redirect('login')
    usuarios = Usuario.objects.all()
    return render(request, 'usuarios/lista.html', {'usuarios': usuarios})


def lista_horarios(request):
    if not request.session.get('usuario_id'):
        return redirect('login')
    horarios = Horario.objects.all()
    return render(request, 'horarios/lista.html', {'horarios': horarios})

def registrar_usuario(request):
    mensaje = None
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            contrasena = form.cleaned_data['contrasena']
            try:
                tipo_usuario = TipoUsuario.objects.get(pk=2)  # Usuario general
                Usuario.objects.create(
                    idTipoUsuario=tipo_usuario,
                    email=email,
                    contrasena=contrasena
                )
                mensaje = "Usuario registrado correctamente. Ahora puedes iniciar sesión."
                form = RegistroUsuarioForm()  # Limpiar formulario
            except IntegrityError:
                mensaje = "Este correo ya está registrado."
    else:
        form = RegistroUsuarioForm()
    return render(request, 'registro.html', {'form': form, 'mensaje': mensaje})

def lista_horarios(request):
    horarios = Horario.objects.select_related('idEmpresa').all()
    return render(request, 'horarios/lista.html', {'horarios': horarios})

def crear_horario(request):
    if request.method == 'POST':
        form = HorarioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_horarios')
    else:
        form = HorarioForm()
    return render(request, 'horarios/formulario.html', {'form': form, 'accion': 'Crear'})

def editar_horario(request, pk):
    horario = get_object_or_404(Horario, pk=pk)
    if request.method == 'POST':
        form = HorarioForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            return redirect('lista_horarios')
    else:
        form = HorarioForm(instance=horario)
    return render(request, 'horarios/formulario.html', {'form': form, 'accion': 'Editar'})

def eliminar_horario(request, pk):
    horario = get_object_or_404(Horario, pk=pk)
    if request.method == 'POST':
        horario.delete()
        return redirect('lista_horarios')
    return render(request, 'horarios/eliminar.html', {'horario': horario})

##rutas:
def lista_rutas(request):
    rutas = Ruta.objects.select_related('idEmpresa').all()
    return render(request, 'rutas/lista.html', {'rutas': rutas})

def crear_ruta(request):
    if request.method == 'POST':
        form = RutaForm(request.POST)
        if form.is_valid():
            ruta = form.save()
            barrios = form.cleaned_data['barrios']
            # Primero borra asociaciones previas (por si acaso)
            RutaBarrio.objects.filter(idRuta=ruta).delete()
            # Crear nuevas asociaciones
            for barrio in barrios:
                RutaBarrio.objects.create(idRuta=ruta, idBarrio=barrio)
            return redirect('lista_rutas')
    else:
        form = RutaForm()
    return render(request, 'rutas/formulario.html', {'form': form})


def editar_ruta(request, pk):
    ruta = get_object_or_404(Ruta, pk=pk)
    if request.method == 'POST':
        form = RutaForm(request.POST, instance=ruta)
        if form.is_valid():
            ruta = form.save()
            barrios = form.cleaned_data['barrios']
            # Actualizar asociaciones
            RutaBarrio.objects.filter(idRuta=ruta).delete()
            for barrio in barrios:
                RutaBarrio.objects.create(idRuta=ruta, idBarrio=barrio)
            return redirect('lista_rutas')
    else:
        # Cargar barrios asociados para que queden seleccionados en el form
        barrios_asociados = Barrio.objects.filter(rutabarrio__idRuta=ruta)
        form = RutaForm(instance=ruta, initial={'barrios': barrios_asociados})
    return render(request, 'rutas/formulario.html', {'form': form, 'ruta': ruta})

def eliminar_ruta(request, pk):
    ruta = get_object_or_404(Ruta, pk=pk)
    if request.method == 'POST':
        ruta.delete()
        return redirect('lista_rutas')
    return render(request, 'rutas/eliminar.html', {'ruta': ruta})

##tiempos
def listar_tiempos(request):
    tiempos = Tiempo.objects.all().order_by('fecha')
    return render(request, 'tiempos/listar.html', {'tiempos': tiempos})

def crear_tiempo(request):
    if request.method == 'POST':
        form = TiempoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_tiempos')
    else:
        form = TiempoForm()
    return render(request, 'tiempos/formulario.html', {'form': form, 'accion': 'Crear'})

def editar_tiempo(request, id):
    tiempo = get_object_or_404(Tiempo, idTiempo=id)
    if request.method == 'POST':
        form = TiempoForm(request.POST, instance=tiempo)
        if form.is_valid():
            form.save()
            return redirect('listar_tiempos')
    else:
        form = TiempoForm(instance=tiempo)
    return render(request, 'tiempos/formulario.html', {'form': form, 'accion': 'Editar'})

def eliminar_tiempo(request, id):
    tiempo = get_object_or_404(Tiempo, idTiempo=id)
    if request.method == 'POST':
        tiempo.delete()
        return redirect('listar_tiempos')
    return render(request, 'tiempos/eliminar.html', {'objeto': tiempo})

##detalle ruta
def lista_detalles_rutas(request):
    detalles = DetalleRuta.objects.select_related('idRuta', 'idTiempo').all().order_by('idRuta_id', 'idTiempo_id')
    return render(request, 'detalle_ruta/lista.html', {'detalles': detalles})

def crear_detalle_ruta(request):
    if request.method == 'POST':
        form = DetalleRutaForm(request.POST)
        if form.is_valid():
            fecha = form.cleaned_data['fecha']
            tiempo_obj, created = Tiempo.objects.get_or_create(fecha=fecha)
            
            detalle_ruta = form.save(commit=False)
            detalle_ruta.idTiempo = tiempo_obj
            detalle_ruta.save()
            return redirect('lista_detalles_rutas')
    else:
        form = DetalleRutaForm()
    return render(request, 'detalle_ruta/crear.html', {'form': form})

def editar_detalle_ruta(request, pk):
    detalle = get_object_or_404(DetalleRuta, pk=pk)
    # Para editar el formulario, pasamos la fecha del tiempo relacionado
    initial = {'fecha': detalle.idTiempo.fecha}
    
    if request.method == 'POST':
        form = DetalleRutaForm(request.POST, instance=detalle)
        if form.is_valid():
            fecha = form.cleaned_data['fecha']
            tiempo_obj, created = Tiempo.objects.get_or_create(fecha=fecha)
            
            detalle = form.save(commit=False)
            detalle.idTiempo = tiempo_obj
            detalle.save()
            return redirect('lista_detalles_rutas')
    else:
        form = DetalleRutaForm(instance=detalle, initial=initial)
    
    return render(request, 'detalle_ruta/editar.html', {'form': form, 'detalle': detalle})

def eliminar_detalle_ruta(request, pk):
    detalle = get_object_or_404(DetalleRuta, pk=pk)
    if request.method == 'POST':
        detalle.delete()
        return redirect('lista_detalles_rutas')
    return render(request, 'detalle_ruta/eliminar.html', {'detalle': detalle})
##barrios
def lista_barrios(request):
    barrios = Barrio.objects.all().prefetch_related('rutabarrio_set__idRuta')
    return render(request, 'barrios/lista.html', {'barrios': barrios})

def crear_barrio(request):
    if request.method == 'POST':
        form = BarrioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_barrios')
    else:
        form = BarrioForm()
    return render(request, 'barrios/form.html', {'form': form, 'accion': 'Crear'})

def editar_barrio(request, pk):
    barrio = get_object_or_404(Barrio, pk=pk)
    if request.method == 'POST':
        form = BarrioForm(request.POST, instance=barrio)
        if form.is_valid():
            form.save()
            return redirect('lista_barrios')
    else:
        form = BarrioForm(instance=barrio)
    return render(request, 'barrios/form.html', {'form': form, 'accion': 'Editar'})

def eliminar_barrio(request, pk):
    barrio = get_object_or_404(Barrio, pk=pk)
    if request.method == 'POST':
        barrio.delete()
        return redirect('lista_barrios')
    return render(request, 'barrios/confirmar_eliminar.html', {'barrio': barrio})

@usuario_autenticado
def actualizar_cuenta(request):
    usuario_id = request.session['usuario_id']
    usuario = Usuario.objects.get(idUsuario=usuario_id)

    if request.method == 'POST':
        form = ActualizarCuentaForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cuenta actualizada correctamente.')
            return redirect('usuario_dashboard')
    else:
        form = ActualizarCuentaForm(instance=usuario)

    return render(request, 'perfil/actualizar_cuenta.html', {'form': form})


@usuario_autenticado
def eliminar_cuenta(request):
    usuario_id = request.session['usuario_id']
    usuario = Usuario.objects.get(idUsuario=usuario_id)

    if request.method == 'POST':
        usuario.delete()
        request.session.flush()  # Cerrar sesión
        messages.success(request, 'Cuenta eliminada exitosamente.')
        return redirect('login')

    return render(request, 'perfil/eliminar_cuenta_confirmar.html')

def buscar_rutas_por_barrio(request):
    barrio_id = request.GET.get('barrio_id')
    rutas_data = []

    if barrio_id:
        rutas = Ruta.objects.filter(rutabarrio__idBarrio=barrio_id).distinct()
        for ruta in rutas:
            rutas_data.append({
                'id': ruta.idRuta,
                'descripcion': str(ruta)  # ejemplo: "Barrio A - Barrio B"
            })

    return JsonResponse({'rutas': rutas_data})

def buscar_rutas_ajax(request):
    barrios = Barrio.objects.all()
    return render(request, 'buscar_rutas_ajax.html', {'barrios': barrios})


def estadisticas_view(request):
    fecha = request.GET.get('fecha')

    # Total de pasajeros por ruta (todos los registros)
    data_total = DetalleRuta.objects.values(
        'idRuta__idEmpresa__nombreEmpresa',
        'idRuta__inicioRuta__nombreBarrio',
        'idRuta__destinoRuta__nombreBarrio'
    ).annotate(total_pasajeros=Sum('cantidadPasajeros'))

    # Media para un día específico (si se selecciona)
    data_media = []
    if fecha:
        data_media = DetalleRuta.objects.filter(idTiempo__fecha=fecha).values(
            'idRuta__idEmpresa__nombreEmpresa',
            'idRuta__inicioRuta__nombreBarrio',
            'idRuta__destinoRuta__nombreBarrio'
        ).annotate(media_pasajeros=Sum('cantidadPasajeros'))

    # Gráfico: Total pasajeros por día y empresa
    resumen = (
        DetalleRuta.objects
        .annotate(fecha=TruncDate('idTiempo__fecha'))
        .values('fecha', 'idRuta__idEmpresa__nombreEmpresa')
        .annotate(total=Sum('cantidadPasajeros'))
        .order_by('fecha')
    )

    fechas = sorted(set(r['fecha'].strftime('%Y-%m-%d') for r in resumen))
    empresas = sorted(set(r['idRuta__idEmpresa__nombreEmpresa'] for r in resumen))

    colores_fijos = ['#36a2eb', '#ff6384', '#ffce56', '#4bc0c0', '#9966ff', '#f67019']

    datasets = []
    for i, empresa in enumerate(empresas):
        data_empresa = []
        for fecha_str in fechas:
            encontrado = next(
                (r['total'] for r in resumen if r['fecha'].strftime('%Y-%m-%d') == fecha_str and r['idRuta__idEmpresa__nombreEmpresa'] == empresa), 0
            )
            data_empresa.append(encontrado)
        datasets.append({
            'label': empresa,
            'data': data_empresa,
            'fill': False,
            'borderColor': colores_fijos[i % len(colores_fijos)],
            'backgroundColor': colores_fijos[i % len(colores_fijos)],
            'tension': 0.1
        })

    # Número de rutas por empresa (para gráfico torta)
    rutas_por_empresa = (
        Ruta.objects.values('idEmpresa__nombreEmpresa')
        .annotate(cantidad=Count('idRuta'))
        .order_by('idEmpresa__nombreEmpresa')
    )
    labels_rutas_empresa = [r['idEmpresa__nombreEmpresa'] for r in rutas_por_empresa]
    data_rutas_empresa = [r['cantidad'] for r in rutas_por_empresa]

    # Total pasajeros por empresa (para gráfico barras)
    pasajeros_por_empresa = (
        DetalleRuta.objects.values('idRuta__idEmpresa__nombreEmpresa')
        .annotate(total_pasajeros=Sum('cantidadPasajeros'))
        .order_by('idRuta__idEmpresa__nombreEmpresa')
    )
    labels_pasajeros_empresa = [p['idRuta__idEmpresa__nombreEmpresa'] for p in pasajeros_por_empresa]
    data_pasajeros_empresa = [p['total_pasajeros'] for p in pasajeros_por_empresa]

    context = {
        'data_total': data_total,
        'data_media': data_media,
        'fecha': fecha,
        'labels_json': json.dumps(fechas),
        'datasets_json': json.dumps(datasets),
        'labels_rutas_empresa': json.dumps(labels_rutas_empresa),
        'data_rutas_empresa': json.dumps(data_rutas_empresa),
        'labels_pasajeros_empresa': json.dumps(labels_pasajeros_empresa),
        'data_pasajeros_empresa': json.dumps(data_pasajeros_empresa),
    }

    return render(request, 'admin/estadisticas.html', context)

def descargar_reporte_csv(request):
    # Crear un libro de Excel en memoria
    wb = openpyxl.Workbook()

    # --- Hoja 1: Detalle ---
    ws1 = wb.active
    ws1.title = "Detalle"

    # Escribir encabezados
    headers_detalle = ['Empresa', 'Inicio', 'Destino', 'Total Pasajeros', 'Media Total Pasajeros']
    ws1.append(headers_detalle)

    # Obtener datos detalle agrupados por ruta
    data = (
        DetalleRuta.objects
        .values(
            'idRuta__idEmpresa__nombreEmpresa',
            'idRuta__inicioRuta__nombreBarrio',
            'idRuta__destinoRuta__nombreBarrio'
        )
        .annotate(total_pasajeros=Sum('cantidadPasajeros'))
        .order_by('-total_pasajeros')
    )

    # Calcular media total global de pasajeros (sobre total_pasajeros)
    total_pasajeros_global = data.aggregate(media_total=Avg('total_pasajeros'))['media_total'] or 0

    # Escribir filas con datos + media global
    for row in data:
        ws1.append([
            row['idRuta__idEmpresa__nombreEmpresa'],
            row['idRuta__inicioRuta__nombreBarrio'],
            row['idRuta__destinoRuta__nombreBarrio'],
            row['total_pasajeros'],
            round(total_pasajeros_global, 2)
        ])

    # --- Hoja 2: Resumen ---
    ws2 = wb.create_sheet(title="Resumen")

    # Encabezados resumen
    headers_resumen = ['Empresa', 'Cantidad Total Pasajeros', 'Media Pasajeros por Ruta', 'Número de Rutas']
    ws2.append(headers_resumen)

    resumen = (
        DetalleRuta.objects
        .values('idRuta__idEmpresa__nombreEmpresa')
        .annotate(
            cantidad_total=Sum('cantidadPasajeros'),
            media_pasajeros=Avg('cantidadPasajeros'),
            num_rutas=Count('idRuta', distinct=True)
        )
        .order_by('idRuta__idEmpresa__nombreEmpresa')
    )

    for row in resumen:
        ws2.append([
            row['idRuta__idEmpresa__nombreEmpresa'],
            row['cantidad_total'],
            round(row['media_pasajeros'] or 0, 2),
            row['num_rutas']
        ])

    # Ajustar ancho columnas para ambas hojas (opcional)
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

    # Guardar en memoria y devolver respuesta HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_pasajeros.xlsx"'
    return response